import csv
import json
import re
import logging
import pandas as pd
import openpyxl
from openpyxl import Workbook,load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# 引入底层的控制器 (假设在同一目录下)
from controller.station_controller import SynthesisStationController
from config.setting import Settings, configure_logging
from config.constants import ResourceCode, TRAY_CODE_DISPLAY_NAME, TraySpec

from driver.exceptions import ValidationError

logger = logging.getLogger("StationManager")

JsonDict = Dict[str, Any]

class StationManager(SynthesisStationController):
    """
    功能:
        上层面向用户的管理器，继承自 SynthesisStationController。
        负责处理 CSV/Excel 文件读取、生成模板，将文件内容转换为中间格式(List/Dict)，
        然后调用父类方法执行具体的业务逻辑。
    """

    def __init__(self, settings: Optional[Settings] = None):
        settings = settings or Settings.from_env()
        configure_logging(settings.log_level)
        super().__init__(settings)

    # ---------- 1. 化合物库文件处理 ----------
    def export_chemical_list_to_file(self, output_path: str) -> None:
        """
        功能:
            获取所有化学品并导出到 CSV 文件
        参数:
            output_path: 输出路径
        返回:
            None
        """
        path = Path(output_path)
        chemical_info = self.get_all_chemical_list()
        chemical_list = chemical_info.get("chemical_list", [])

        if not chemical_list:
            logger.warning("化学品列表为空，未写入文件")
            return

        fieldnames = [
            "fid", "name", "sssi", "cas", "element", "state",
            "concentration_str", "chemical_properties", "preparation_method"
        ]
        
        # 确保目录存在
        path.parent.mkdir(parents=True, exist_ok=True)

        with path.open("w", newline="", encoding="utf-8-sig") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for item in chemical_list:
                writer.writerow(item)
        
        logger.info(f"化学品列表已导出至: {path.resolve()}")

    def sync_chemicals_from_file(self, file_path: str, overwrite: bool = False) -> None:
        """
        功能:
            读取 CSV 文件并通过父类同步化学品到工站
        参数:
            file_path: CSV 文件路径
            overwrite: 是否覆盖更新
        返回:
            None
        """
        path = Path(file_path)
        if not path.exists():
            # 生成模板
            header = ["name", "cas", "element", "state", "concentration_str", "chemical_properties", "preparation_method"]
            with path.open("w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerow(header)
            logger.warning(f"文件不存在，已生成模板: {path}")
            return

        # 读取并清洗数据
        items: List[JsonDict] = []
        with path.open("r", newline="", encoding="utf-8-sig") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                name = (row.get("name") or "").strip()
                state = (row.get("state") or "").strip()
                if name and state:
                    # 过滤空值键
                    clean_item = {k: v.strip() for k, v in row.items() if v and str(v).strip()}
                    items.append(clean_item)
        
        # 调用父类逻辑处理
        self.sync_chemicals_from_data(items, overwrite=overwrite)

    def check_chemical_library_by_file(self, file_path: str) -> Dict[str, List[str]]:
        """
        功能:
            读取化学品库文件并调用底层校验逻辑，输出校验结果
        参数:
            file_path: str, 化学品库文件路径，支持 Excel/CSV
        返回:
            Dict[str, List[str]], 包含 errors 与 warnings
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {path}")

        # 读取文件后交给控制层做校验
        df = pd.read_excel(path) if path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(path)
        df = df.fillna("")
        rows = df.to_dict(orient="records")
        headers = [str(col).strip() for col in df.columns]

        result = self.check_chemical_library_data(rows, headers)

        for msg in result.get("warnings", []):
            logger.warning(msg)

        if len(result.get("errors", [])) > 0:
            for msg in result["errors"]:
                logger.error(msg)
            raise ValidationError("化学品库完整性检查未通过，请修复错误后重试")

        return result
    
    def deduplicate_chemical_library_by_file(self, file_path: str, output_path: Optional[str] = None) -> List[JsonDict]:
        """
        功能:
            读取化学品库文件，按 substance 自动去重并回写
        参数:
            file_path: str, 输入文件路径，支持 Excel/CSV
            output_path: Optional[str], 输出文件路径，默认覆盖原文件
        返回:
            List[Dict[str, Any]], 去重后的数据
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {path}")

        df = pd.read_excel(path) if path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(path)
        df = df.fillna("")
        headers = [str(c).strip() for c in df.columns]
        rows = df.to_dict(orient="records")

        dedup_rows = self.deduplicate_chemical_library_data(rows, headers)

        target_path = Path(output_path) if output_path else path
        out_df = pd.DataFrame(dedup_rows)
        if target_path.suffix.lower() == ".csv":
            out_df.to_csv(target_path, index=False, encoding="utf-8-sig")
        else:
            out_df.to_excel(target_path, index=False)
            self._beautify_excel_database(target_path)  # 保存后再美化

        logger.info("化合物库去重完成，输出文件: %s", target_path.resolve())
        return dedup_rows
    
    def _beautify_excel_database(self, file_path: Path) -> None:
        """
        功能:
            美化去重后的 Excel: 表头加粗、全居中、列宽自适应、按内容选择中英文字体
        参数:
            file_path: Path, 目标 Excel 路径
        返回:
            None
        """
        wb = load_workbook(file_path)
        ws = wb.active
        MAX_WIDTH = 60  # 列宽上限

        align_center = Alignment(horizontal="center", vertical="center")

        def _is_chinese(text: str) -> bool:
            return re.search(r"[\u4e00-\u9fff]", text) is not None

        # 遍历列计算列宽并设置字体/对齐
        for col_cells in ws.iter_cols():
            max_len = 0
            for idx, cell in enumerate(col_cells):
                val_str = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val_str))

                # 按内容切换字体，表头加粗
                if idx == 0:
                    cell.font = Font(name="宋体", bold=True) if _is_chinese(val_str) else Font(name="Arial", bold=True)
                else:
                    cell.font = Font(name="宋体") if _is_chinese(val_str) else Font(name="Arial")

                cell.alignment = align_center

            # 列宽留一点边距，最小 10，最大 40
            col_width = max(10, max_len + 2)
            col_width = min(col_width, MAX_WIDTH)
            ws.column_dimensions[col_cells[0].column_letter].width = col_width

        wb.save(file_path)

    def align_chemicals_with_file(self, file_path: str, auto_delete: bool = True) -> None:
        """
        功能:
            读取 Excel/CSV 文件，调用父类对齐逻辑，并将结果(fid)写回文件
        参数:
            file_path: 文件路径
            auto_delete: 是否删除不在文件中的工站化学品
        返回:
            None
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"未找到化学品对齐文件: {path}")

        # 读取文件内容为 List[Dict]
        df = pd.read_excel(path) if path.suffix in ['.xlsx', '.xls'] else pd.read_csv(path)
        # 将 NaN 替换为空字符串
        df = df.fillna("")
        rows = df.to_dict(orient='records')
        header = df.columns.tolist()

        # 调用父类进行对齐，父类会修改 rows 中的数据(如回填 chemical_id)
        updated_rows = self.align_chemicals_from_data(rows, auto_delete=auto_delete)

        # 写回文件
        new_df = pd.DataFrame(updated_rows)
        # 保持原有列顺序，如果增加了新列(如 chemical_id 之前没有)，这会包含它
        if path.suffix == '.csv':
            new_df.to_csv(path, index=False, encoding="utf-8-sig")
        else:
            new_df.to_excel(path, index=False)
            self._beautify_excel_database(path)  # 保存后再美化
        
        logger.info(f"化学品对齐完成并回写文件: {path}")

    # ---------- 2. 上料文件处理 ----------
    def batch_in_tray_by_file(self, file_path: str) -> JsonDict:
        """
        功能:
            读取上料表格，转换为中间格式，调用父类生成 Payload 并执行上料
        参数:
            file_path: 文件路径
        返回:
            Dict: API 响应
        """
        path = Path(file_path)
        if not path.exists():
            logger.warning(f"未找到{file_path}.自动生成模板文件")
            self._generate_batch_in_tray_template(path.with_suffix(".xlsx"))
            return {}

        rows: List[Tuple[str, str, str]] = []
        
        # 读取文件
        if path.suffix == '.xlsx':
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 确保取前三列，且处理 None
                pos = str(row[0]) if row[0] is not None else ""
                t_type = str(row[1]) if len(row) > 1 and row[1] is not None else ""
                content = str(row[2]) if len(row) > 2 and row[2] is not None else ""
                rows.append((pos, t_type, content))
        else:
            df = pd.read_csv(path)
            df = df.fillna("")
            for _, row in df.iterrows():
                rows.append((str(row[0]), str(row[1]), str(row[2])))

        # 调用父类生成 Payload
        payload = self.build_batch_in_tray_payload(rows)
        
        if not payload:
            logger.warning("生成的上料数据为空")
            return {}
            
        # 执行上料
        return payload
        # return self.batch_in_tray(payload)

    def _generate_batch_in_tray_template(self, file_path: Path) -> None:
        """
        功能:
            生成批量上料Excel模板, 配置上料点位下拉、托盘类型下拉与内容示例
        参数:
            file_path: Path, 模板输出路径
        返回:
            None
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "batch_in_tray"
        ws.append(["position", "tray_type", "content"])
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 80

        # 位置下拉，包含 TB 列与 W-1-1~W-1-8 货位
        positions_tb = [f"TB-{row}-{col}" for row in (1, 2) for col in range(1, 5)]
        positions_w = [f"W-1-{index}" for index in range(1, 9)]
        positions = positions_tb + positions_w
        dv_pos = DataValidation(type="list", formula1=f"\"{','.join(positions)}\"")
        ws.add_data_validation(dv_pos)
        dv_pos.add("A2:A101")

        # 托盘下拉，耗材显示数量范围，带物质显示点位范围
        consumable_trays = {
            int(ResourceCode.TIP_TRAY_50UL),
            int(ResourceCode.TIP_TRAY_1ML),
            int(ResourceCode.TIP_TRAY_5ML),
            int(ResourceCode.REACTION_SEAL_CAP_TRAY),
            int(ResourceCode.FLASH_FILTER_INNER_BOTTLE_TRAY),
            int(ResourceCode.FLASH_FILTER_OUTER_BOTTLE_TRAY),
            int(ResourceCode.REACTION_TUBE_TRAY_2ML),
            int(ResourceCode.TEST_TUBE_MAGNET_TRAY_2ML),
        }
        tray_display: List[str] = []
        for code, name in TRAY_CODE_DISPLAY_NAME.items():
            base_text = f"{name}({code})"
            try:
                enum_name = ResourceCode(code).name
                spec = getattr(TraySpec, enum_name, None)
            except Exception:
                spec = None

            if spec is None:
                tray_display.append(base_text)
                continue

            col_count, row_count = spec
            if col_count <= 0 or row_count <= 0:
                tray_display.append(base_text)
                continue

            if code in consumable_trays:
                capacity = col_count * row_count
                tray_display.append(f"{base_text} [1-{capacity}]")
            else:
                end_row_char = chr(ord("A") + row_count - 1)
                tray_display.append(f"{base_text} [A1-{end_row_char}{col_count}]")

        # 用隐藏sheet作为数据源，避免下拉字符串过长
        tray_sheet = wb.create_sheet("validation_meta")
        for idx, option in enumerate(tray_display, start=1):
            tray_sheet.cell(row=idx, column=1).value = option
        tray_sheet.sheet_state = "hidden"

        dv_tray = DataValidation(
            type="list",
            formula1=f"=validation_meta!$A$1:$A${len(tray_display)}",
            showInputMessage=True,
        )
        ws.add_data_validation(dv_tray)
        dv_tray.add("B2:B101")

        ws["C1"] = "content(耗材填数量; 物质填: A1|名称|2mL; B2|名称|5mg)"
        wb.save(file_path)
        logger.info(f"已生成上料模板: {file_path}")

    # ---------- 3. 任务生成文件处理 ----------
    def create_task_by_file(self, template_path: str, chemical_db_path: str) -> JsonDict:
        """
        功能:
            读取任务模板和化学品库，解析为中间数据，调用父类生成任务 Payload 并提交
        参数:
            template_path: 实验模板路径
            chemical_db_path: 化学品库路径
        返回:
            Dict: 任务创建结果
        """
        t_path = Path(template_path)
        c_path = Path(chemical_db_path)

        # 1. 检查并生成模板
        if not t_path.exists():
            self._generate_reaction_template(t_path)
            raise FileNotFoundError(f"已生成模板 {t_path}，请填写后重试")

        if not c_path.exists():
            raise FileNotFoundError(f"未找到化学品库文件: {c_path}")

        # 2. 读取化学品库 -> Dict
        chem_df = pd.read_excel(c_path) if c_path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(c_path)
        chem_df.columns = [str(c).strip().lower() for c in chem_df.columns]

        def _pick(row, *keys, default=None):
            for k in keys:
                if k in row and pd.notna(row[k]):
                    return row[k]
            return default

        chemical_db: Dict[str, Dict[str, Any]] = {}
        for _, r in chem_df.iterrows():
            row = {k: r.get(k) for k in chem_df.columns}
            name = str(_pick(row, "substance", "name", "chemical_name", default="") or "").strip()
            if not name:
                continue

            chemical_db[name] = {
                "chemical_id": _pick(row, "chemical_id"),
                "molecular_weight": _pick(row, "molecular_weight", "mw"),
                "physical_state": str(_pick(row, "physical_state", "state", default="") or "").strip().lower(),
                # 统一把各种写法都接住（你原先 lower 后再用 'density (g/mL)' 是取不到的）
                "density (g/mL)": _pick(
                    row,
                    "density (g/ml)",
                    "density(g/ml)",
                    "density_g_ml",
                    "density",
                    default=None,
                ),
                "fid": _pick(row, "fid"),
            }

        # 3. 读取任务模板 -> params(Dict), headers(List), data_rows(List[List])
        wb = load_workbook(t_path, data_only=True)
        ws = wb.active

        # 3.1 找到表头行/实验编号列（模板里一般是：row=1, col=3）
        header_row = None
        exp_no_col = None
        for r in range(1, min(ws.max_row, 50) + 1):
            for c in range(1, min(ws.max_column, 50) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and "实验编号" in v:
                    header_row, exp_no_col = r, c
                    break
            if header_row is not None:
                break
        if header_row is None or exp_no_col is None:
            raise ValueError("模板中未找到'实验编号'表头")

        # 3.2 提取全局参数（左侧 A/B）
        # - 实验名称：A1是标签，用户通常填在 B1
        params: Dict[str, Any] = {}
        exp_name = ws.cell(1, 2).value  # B1
        if exp_name is not None and str(exp_name).strip() != "":
            params["实验名称"] = str(exp_name).strip()

        # 扫描 A/B（从第2行开始，遇到“注：”不停止也可以；这里仅跳过“注：”本行）
        for r in range(2, ws.max_row + 1):
            key = ws.cell(r, 1).value
            val = ws.cell(r, 2).value

            if key is None:
                continue
            key_str = str(key).strip()
            if not key_str:
                continue

            # 跳过注释行（不写入 params；否则会污染）
            if key_str.startswith("注：") or key_str.startswith("注:"):
                continue

            # 分类标题行通常是合并单元格，B 为空；这类不要写入 params
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue

            params[key_str] = val

        # 3.3 生成 headers（从 “实验编号”列开始往右：C..M）
        # 同时把 “试剂_1” -> “试剂名称_1”，让 build_task_payload 能识别
        raw_headers: List[Any] = []
        for c in range(exp_no_col, ws.max_column + 1):
            raw_headers.append(ws.cell(header_row, c).value)

        headers: List[str] = []
        reagent_idx = 0
        for h in raw_headers:
            s = "" if h is None else str(h).strip()

            # 规范化：试剂_1/试剂1 -> 试剂名称_1
            if s.startswith("试剂") and "量" not in s and s != "试剂名称":
                reagent_idx += 1
                headers.append(f"试剂名称_{reagent_idx}")
                continue

            # 规范化：试剂量 -> 试剂量_1/2/...
            if "试剂量" in s:
                # 若前面还没遇到试剂列，给个兜底编号
                idx = reagent_idx if reagent_idx > 0 else (len([x for x in headers if "试剂量" in x]) + 1)
                headers.append(f"试剂量_{idx}")
                continue

            headers.append(s)

        # 3.4 生成 data_rows：从表头下一行开始，按实验编号列读取到最后一列（C..M）
        data_rows: List[List[Any]] = []
        for r in range(header_row + 1, ws.max_row + 1):
            exp_no = ws.cell(r, exp_no_col).value

            # 实验编号为空：认为实验区结束（模板一般后面都是空）
            if exp_no is None or (isinstance(exp_no, str) and exp_no.strip() == ""):
                # 只有在已经读到至少一行实验后才 break，避免中间空行误判
                if data_rows:
                    break
                else:
                    continue

            row_vals: List[Any] = []
            for c in range(exp_no_col, ws.max_column + 1):
                v = ws.cell(r, c).value
                # 这里不要强制 str 化，build_task_payload 内部会 str()；但 None 要变成 ""
                row_vals.append("" if v is None else v)

            data_rows.append(row_vals)

        # 4. 调用父类纯逻辑生成 Payload
        task_payload = self.build_task_payload(params, headers, data_rows, chemical_db)

        # 5. 提交任务信息到工站
        resp = self.add_task(task_payload)

        self._assert_success(resp,"创建任务")

        # 6. 提交任务信息到工站
        task_id = resp.get("task_id")

        return task_id

    def _generate_reaction_template(self, path: Path) -> None:
        """
        生成与 reeaction_template.xlsx 一致的反应模板
        结构：左侧为参数配置区，右侧为实验试剂填报区
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 模板默认字体：等线 11
        base_font = Font(name="等线", charset=134, family=2, scheme="minor", sz=11)
        title_font = Font(name="等线", charset=134, family=2, scheme="minor", sz=11, bold=True)
        center = Alignment(horizontal="center", vertical="center")

        # --- 1. 定义左侧参数配置数据 (行2开始, A列和B列) ---
        left_params = [
            ("反应设定", ""),
            ("实验名称", "Auto_task"),
            ("反应器类型", "heat"),
            ("反应时间(h)", 8),
            ("反应温度(°C)", 40),
            ("转速(rpm)", 500),
            ("搅拌后⽬标温度(°C)", 30),
            ("等待目标温度", "否"),
            ("称量设定", ""),
            ("称量误差(%)", 3),
            ("最大称量误差(mg)", 1),
            ("加料设定", ""),
            ("固定加料顺序", "否"),
            ("自动加磁子", "是"),
            ("内标设定", ""),
            ("内标种类", "1,3,5-三异丙基苯(内标,1mol/L in MeCN)"),
            ("内标用量(μL/mg)", 100),
            ("稀释液种类", "乙腈"),
            ("稀释量(μL)", 500),
            ("取样量(μL)", 2),
            ("加入内标后搅拌时间(min)", 5),
            ("", ""),  # 空行
        ]

        # --- 2. 设置第一行表头 (Row 1) ---
        ws.cell(row=1, column=3, value="实验编号").font = base_font
        
        reagent_count = 5
        current_col = 4
        for i in range(1, reagent_count + 1):
            ws.cell(row=1, column=current_col, value=f"试剂").font = base_font
            ws.cell(row=1, column=current_col + 1, value="试剂量").font = base_font
            current_col += 2

        # --- 3. 填充左侧参数区 (Row 2 ~ Row 22) ---
        for idx, (param_name, default_val) in enumerate(left_params):
            row_idx = idx + 1  # 从第2行开始

            # 分类标题：模板是 A:B 合并，只写 A 列，且加粗
            if param_name and default_val == "":
                ws.cell(row=row_idx, column=1, value=param_name).font = title_font
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                continue

            # 空行：保持空
            if param_name == "" and default_val == "":
                continue

            # 普通参数行
            ws.cell(row=row_idx, column=1, value=param_name).font = base_font
            ws.cell(row=row_idx, column=2, value=default_val).font = base_font

        # --- 4. 填充右侧实验编号 (Row 2 ~ Row 25) ---
        for i in range(1, 25):  # 1~24
            row_idx = i + 1
            ws.cell(row=row_idx, column=3, value=i).font = base_font

        # --- 5. 底部注释 (Row 23) ---
        note_row = len(left_params) + 2  # 23
        note_text = "注：试剂量支持单位：(mmol,g,mg,μL,mL）"
        ws.cell(row=note_row, column=1, value=note_text).font = base_font
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
        ws.cell(row=note_row, column=1).alignment = center  # 合并后的单元格居中

        # --- 6. 字体铺满 (A1:M25)  ---
        for r in range(1, 26):
            for c in range(1, 14):  # A..M
                cell = ws.cell(r, c)
                # 标题行的粗体不要覆盖
                if cell.font and cell.font.bold:
                    continue
                cell.font = base_font

        # --- 7. 对齐 ---
        # C~L 整块都居中（含空白）
        for r in range(1, 26):
            for c in range(3, 13):  # C..L
                ws.cell(r, c).alignment = center

        # A 列：1~21 + 23 行居中（22/24/25 行保持默认）
        for r in list(range(1, 22)) + [23]:
            ws.cell(r, 1).alignment = center

        # B 列：只有有值的参数行居中（标题行/空白行/合并后的 B 不处理）
        for r in [2, 3, 4, 5, 6, 7, 8, 10, 11, 13, 14, 16, 17, 18, 19, 20, 21]:
            ws.cell(r, 2).alignment = center

        # M 列：只有表头 M1 居中
        ws.cell(1, 13).alignment = center

        # 表头 A1/C1 也居中（模板如此）
        ws.cell(1, 1).alignment = center
        ws.cell(1, 3).alignment = center

        # --- 8. 列宽： ---
        widths_map = {
            "A": 26.0,
            "B": 38.0,
            "C": 15.0,
            "D": 14.0,
            "E": 14.0,
            "F": 14.0,
            "G": 14.0,
            "H": 14.0,
            "I": 14.0,
            "J": 14.0,
            "K": 14.0,
            "L": 14.0,
            "M": 14.0,
        }
        for col_letter, w in widths_map.items():
            ws.column_dimensions[col_letter].width = w

        wb.save(path)
        logger.info(f"已生成任务模板: {path}")

    # ---------- 4. 任务物料和站内物料对比 ----------


if __name__ == "__main__":

    # 测试代码
    try:
        settings = Settings.from_env()
        manager = StationManager(settings)

        #---------------提交任务流程-------------------

        # 0. 设定文件名称

        # 提交任务文件
        task_tpl = Path("reaction_template_5.xlsx")

        # 化合物库文件
        chem_db = Path("chemical_list.xlsx")

        # 进料文件
        template_in = Path("batch_in_tray.xlsx")

        # 3. 本地化学品库去重整理
        # manager.deduplicate_chemical_library_by_file(chem_db)

        # 4. 本地化学品库数据完整性检验
        # manager.check_chemical_library_by_file(chem_db)

        # 5. 工站化学品库和本地化学品库数据对齐
        # manager.align_chemicals_with_file(chem_db)

        # 6. 上传任务到工站
        # task_id = manager.create_task_by_file(str(task_tpl), str(chem_db))

        # 4. 对比站内资源和任务文件json列出缺乏
        # manager.check_resource_for_task(str(task_tpl), str(chem_db))

        # 5. 上料
        # manager.batch_in_tray_by_file(str(template_in))

        # # 7 . 开始任务
        # resp = manager.start_task(task_id)

        #---------------工站状态查询-------------------

        # 1. 查询站内所有物料信息
        # resource_info = manager.get_resource_info()

        # 2. 查询站内所有设设备状态
        devices_info = manager.list_device_status()
        print(devices_info)

        # 3. 查询工站运行状态
        station_info = manager.station_state()

        print(station_info)
        # 4. 查询手套箱状态
        glovebox_info = manager.get_glovebox_env()
        print(glovebox_info)

        #---------------其他可执行动作-------------------

        # 1. 登录
        # manager.login()

        # 2. 设备初始化
        # manager.device_init()

        
    except Exception as e:
        logger.error(f"测试失败: {e}", exc_info=True)

        #————————————————额外功能————————————————————

        # # 获取站内所有化学品信息,导出到csv文件
        # manager.export_chemical_list_to_file("chemicals_list_export.csv")

        # # 通过csv进行化学品录入
        # manager.sync_chemicals_from_file("add_chemical_list.csv")
